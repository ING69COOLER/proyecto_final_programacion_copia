# excel_exporter.py
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from openpyxl.cell.cell import MergedCell

class ExcelExporter:
    def __init__(self, lista_eventos, lista_personas, archivo_excel):
        self.lista_eventos = lista_eventos
        self.lista_personas = lista_personas
        self.archivo_excel = archivo_excel

    def delete_existing_file(self):
        """Elimina el archivo Excel existente si ya existe."""
        if os.path.exists(self.archivo_excel):
            os.remove(self.archivo_excel)
            print(f"Archivo existente '{self.archivo_excel}' eliminado.")

    def create_excel_file(self):
        """Crea un archivo Excel con una hoja para cada evento y un gráfico de barras."""
        workbook = Workbook()
        del workbook["Sheet"]

        eventos_df = pd.DataFrame(self.lista_eventos)
        personas_df = pd.DataFrame(self.lista_personas)

        if 'idEvento' not in personas_df.columns:
            print("Error: La columna 'idEvento' no existe en los datos de personas.")
            return

        for _, evento in eventos_df.iterrows():
            evento_id = evento['id']
            evento_nombre = evento.get('nombre', f"Evento_{evento_id}")
            personas_evento_df = personas_df[personas_df['idEvento'] == evento_id]

            if not personas_evento_df.empty:
                sheet = workbook.create_sheet(title=str(evento_nombre))
                self.populate_sheet(sheet, personas_evento_df, evento_nombre)
                self.add_chart(sheet, evento_nombre)

        workbook.save(self.archivo_excel)
        print(f"Exportación completada: {self.archivo_excel} creado.")

    def populate_sheet(self, sheet, data_df, evento_nombre):
        """Llena una hoja de Excel con los datos de un DataFrame y aplica estilos."""
        # Título del evento en la hoja
        title_cell = sheet.cell(row=1, column=1, value=f"Detalles del Evento: {evento_nombre}")
        title_cell.font = Font(bold=True, size=14, color="4F81BD")
        title_cell.alignment = Alignment(horizontal="center")
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

        # Encabezado y datos del DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(data_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = sheet.cell(row=r_idx, column=c_idx, value=value)
                self.apply_styles(cell, is_header=(r_idx == 3))

        self.adjust_column_width(sheet)

    def add_chart(self, sheet, evento_nombre):
        """Agrega un gráfico de barras para comparar los ingresos entre sillas VIP y regulares."""
        # Crear un DataFrame a partir de los valores en la hoja actual

        data_df = pd.DataFrame(sheet.values)
        data_df.columns = data_df.iloc[2]  # Usar la tercera fila como encabezados de columna
        data_df = data_df.drop([0, 1, 2])  # Eliminar las primeras filas innecesarias

        # Asegurarse de que los datos de "totalPagar" sean numéricos
        data_df["totalPagar"] = pd.to_numeric(data_df["totalPagar"], errors="coerce")

        data_df = data_df[data_df["tipoSilla"].isin(["sillas_vip", "sillas_regular"])]
        

        
        # Agrupar los datos por tipo de silla y sumar los ingresos
        ingresos_df = data_df.groupby("tipoSilla", as_index=False)["totalPagar"].sum()

        # Asegurar que ambos tipos de silla estén presentes en el DataFrame para el gráfico
        for tipo in ["sillas_vip", "sillas_regular"]:
            if tipo not in ingresos_df["tipoSilla"].values:
                ingresos_df = pd.concat([ingresos_df, pd.DataFrame({"tipoSilla": [tipo], "totalPagar": [0]})], ignore_index=True)

        # Agregar los datos de ingresos por tipo de silla a la hoja de Excel para el gráfico
        start_row = sheet.max_row + 3  # Añadir datos después de los datos existentes
        sheet.cell(row=start_row, column=1, value="Tipo de Silla").font = Font(bold=True)
        sheet.cell(row=start_row, column=2, value="Total Recaudado").font = Font(bold=True)

        for i, (tipo, total) in enumerate(ingresos_df.values, start=1):
            sheet.cell(row=start_row + i, column=1, value=tipo)
            sheet.cell(row=start_row + i, column=2, value=total)

        # Crear el gráfico de barras
        chart = BarChart()
        chart.title = f"Ingresos Comparativos por Tipo de Silla - {evento_nombre}"
        chart.x_axis.title = "Tipo de Silla"
        chart.y_axis.title = "Ingresos Totales"
        chart.style = 10
        chart.varyColors = True

        # Crear referencias de datos para el gráfico
        data = Reference(sheet, min_col=2, min_row=start_row, max_row=start_row + len(ingresos_df))
        cats = Reference(sheet, min_col=1, min_row=start_row + 1, max_row=start_row + len(ingresos_df))
        
        # Agregar los datos y categorías al gráfico
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Insertar el gráfico en la hoja
        chart_position = f"E{start_row}"  # Ajustar la posición del gráfico según sea necesario
        sheet.add_chart(chart, chart_position)

    @staticmethod
    def apply_styles(cell, is_header=False):
        """Aplica estilos a una celda."""
        if is_header:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill("solid", fgColor="4F81BD")
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
        cell.border = thin_border

    @staticmethod
    def adjust_column_width(sheet):
        """Ajusta el ancho de las columnas para adaptarse al contenido, omitiendo celdas fusionadas."""
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = None
            
            for cell in column_cells:
                if column_letter is None and not isinstance(cell, MergedCell):
                    column_letter = cell.column_letter

                if not isinstance(cell, MergedCell) and cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            if column_letter:
                sheet.column_dimensions[column_letter].width = max_length + 2
